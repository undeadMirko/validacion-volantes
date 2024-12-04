from PyQt5.QtWidgets import (
    QMainWindow, QVBoxLayout, QLabel, QPushButton, QWidget, 
    QFormLayout, QDateEdit, QTextEdit, QMessageBox
)
from PyQt5.QtCore import QDate
from workers.selenium_worker import SeleniumWorker
import requests
import json
from bs4 import BeautifulSoup
from openpyxl import Workbook 
from openpyxl.styles import *
import config

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Validación de Volantes")
        self.setGeometry(100, 100, 800, 600)

        # Diseño de la ventana principal
        main_layout = QVBoxLayout()

        # Campo para la fecha de consulta
        form_layout = QFormLayout()
        self.fecha_label = QLabel("Fecha de Consulta:")
        self.fecha_input = QDateEdit()
        self.fecha_input.setDate(QDate.currentDate())
        form_layout.addRow(self.fecha_label, self.fecha_input)
        main_layout.addLayout(form_layout)

        # Botones principales
        self.entrar_poliedro_button = QPushButton("Entrar a Poliedro")
        self.entrar_poliedro_button.clicked.connect(self.entrar_poliedro)

        self.hacer_consulta_button = QPushButton("Hacer Consulta")
        self.hacer_consulta_button.clicked.connect(self.hacer_consulta)

        self.exportar_excel_button = QPushButton("Exportar a Excel")
        self.exportar_excel_button.clicked.connect(self.exportar_a_excel)

        self.limpiar_button = QPushButton("Limpiar Información")
        self.limpiar_button.clicked.connect(self.limpiar_informacion)

        main_layout.addWidget(self.entrar_poliedro_button)
        main_layout.addWidget(self.hacer_consulta_button)
        main_layout.addWidget(self.exportar_excel_button)
        main_layout.addWidget(self.limpiar_button)

        # Área para ver el progreso del proceso
        self.progreso_text = QTextEdit()
        self.progreso_text.setReadOnly(True)
        main_layout.addWidget(QLabel("Progreso del Proceso:"))
        main_layout.addWidget(self.progreso_text)

        # Contenedor principal
        container = QWidget()
        container.setLayout(main_layout)
        self.setCentralWidget(container)

        # Variables
        self.current_worker = None
        self.cookie = None
        self.resultados = []

    def log_progreso(self, mensaje):
        """Escribe un mensaje en el área de progreso."""
        self.progreso_text.append(mensaje)

    def entrar_poliedro(self):
        """Inicia sesión en Poliedro y habilita la monitorización de solicitudes."""
        self.log_progreso("Abriendo Poliedro. Por favor, inicie sesión manualmente y realice una consulta.")
        self.current_worker = SeleniumWorker()
        self.current_worker.cookie_capturado.connect(self.mostrar_header)
        self.current_worker.start()

    def mostrar_header(self, cookie):
        """Muestra el cookie capturado."""
        self.cookie = cookie
        self.log_progreso(f"Cookie capturada: {self.cookie}")

    def hacer_consulta(self):
        """Realiza consultas automáticamente para cada código pendiente con Volantetype R y V."""
        if not self.cookie:
            self.log_progreso("Por favor, captura la cookie antes de realizar una consulta.")
            return

        # Formatear la fecha seleccionada
        fecha = self.fecha_input.date().toString("dd/MM/yyyy")

        headers = {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "Accept-Encoding": "gzip, deflate, br, zstd",
            "Accept-Language": "es-US,es-419;q=0.9,es;q=0.8",
            "Cache-Control": "max-age=0",
            "Connection": "keep-alive",
            "Content-Type": "application/x-www-form-urlencoded",
            "Referer": "https://poliedrodist.comcel.com.co/Recaudo.PS/VolantesNIT/Index",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "same-origin",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
            "sec-ch-ua": '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"Windows"',
            "Cookie": self.cookie,
        }

        url = "https://poliedrodist.comcel.com.co/Recaudo.PS/VolantesNIT/ResumenVolantes"

        for codigo, items in config.codigos_pendientes.items():
            for item in items:
                for volantetype in ['R', 'V']:
                    data = {
                        "SelectedDistr": f"-{item}",
                        "Volantetype": volantetype,
                        "Initialdate": fecha,
                        "FinalDate": fecha,
                    }

                    self.log_progreso(f"Haciendo consulta para: {item} con Volantetype {volantetype}...")

                    try:
                        response = requests.post(url, headers=headers, data=data)
                        if response.status_code == 200:
                            soup = BeautifulSoup(response.text, 'html.parser')
                            if soup.find('form', {'id': 'frmConsultaVolantes'}):
                                table = soup.find('form', {'id': 'frmConsultaVolantes'}).find('table')
                                rows = []
                                for row in table.find_all('tr')[1:]:
                                    cols = row.find_all('td')
                                    if len(cols) >= 3:
                                        secuencia = cols[0].get_text(strip=True)
                                        valor_efectivo = cols[1].get_text(strip=True)
                                        valor_cheque = cols[2].get_text(strip=True)
                                        rows.append({
                                            "Secuencia": secuencia,
                                            "Valor Efectivo": valor_efectivo,
                                            "Valor Cheque": valor_cheque,
                                        })
                                self.resultados.append({
                                    "Codigo": codigo,
                                    "Distribuidor": item,
                                    "Fecha": fecha,
                                    "Volantetype": volantetype,
                                    "Datos": rows
                                })
                            elif soup.find('div', {'id': 'MessageSinReg'}):
                                self.resultados.append({
                                    "Codigo": codigo,
                                    "Distribuidor": item,
                                    "Fecha": fecha,
                                    "Volantetype": volantetype,
                                    "Mensaje": "No existe ningún Volante asociado a los filtros de búsqueda."
                                })
                            else:
                                self.resultados.append({
                                    "Codigo": codigo,
                                    "Distribuidor": item,
                                    "Fecha": fecha,
                                    "Volantetype": volantetype,
                                    "Mensaje": "Respuesta inesperada, no se pudo procesar el HTML correctamente."
                                })

                            self.log_progreso(f"Consulta exitosa para: {item} con Volantetype {volantetype}.")
                        else:
                            self.log_progreso(f"Error en consulta para {item} con Volantetype {volantetype}: {response.status_code}. Respuesta: {response.text}")
                    except Exception as e:
                        self.log_progreso(f"Error durante la consulta para {item} con Volantetype {volantetype}: {e}")

        self.log_progreso("Consultas finalizadas.")
        self.mostrar_resultados()
        self.mostrar_alerta("Consulta Finalizada", "Las consultas se han completado correctamente.")

    def mostrar_alerta(self, titulo, mensaje):
        """Muestra un mensaje de alerta."""
        QMessageBox.information(self, titulo, mensaje)

    def limpiar_informacion(self):
        """Limpia los campos de la interfaz pero conserva la cookie."""
        self.log_progreso("Limpiando campos...")
        
        # Limpiar los campos visibles pero conservar la cookie
        self.fecha_input.clear()  # Limpia el campo de fecha
        self.progreso_text.clear()  # Limpia el área de progreso
        
        # La cookie se conserva porque está almacenada en self.cookie
        # No es necesario modificarla, solo limpiar la interfaz
        self.log_progreso(f"Cookie conservada: {self.cookie}")

    def mostrar_resultados(self):
        """Muestra los resultados en la interfaz."""
        if not self.resultados:
            self.log_progreso("No se han encontrado resultados.")
            return

        for resultado in self.resultados:
            self.log_progreso(f"Resultado para {resultado['Distribuidor']} ({resultado['Volantetype']}): {json.dumps(resultado, indent=4)}")

    from openpyxl.styles import Alignment, Border, Side, PatternFill, Font

    def exportar_a_excel(self):
        """Exporta los resultados a un archivo Excel."""
        if not self.resultados:
            self.log_progreso("No hay resultados para exportar.")
            return

        distribuidor_a_codigo = {
            "11855": "D2146.00001", "11894": "D2146.00002", "11896": "D2146.00003", "11898": "D2146.00004", 
            "12498": "D2146.00005", "14819": "D2146.00006", "15002": "D2146.00007", "18249": "D2146.00008", 
            "23090": "D2146.00010", "44914": "D2146.00011", "44916": "D2146.00012", "44920": "D2146.00014", 
            "54542": "D2146.00024", "54554": "D2146.00027", "56084": "D2146.00028", "60478": "D2146.00029", 
            "63827": "D2146.00030", "63831": "D2146.00031", "63834": "D2146.00032", "63837": "D2146.00033", 
            "63839": "D2146.00034", "63841": "D2146.00035", "63843": "D2146.00036", "68064": "D2146.00037", 
            "68068": "D2146.00038", "68072": "D2146.00039", "68076": "D2146.00040", "69026": "D2146.00041", 
            "70037": "D2146.00042", "71332": "D2146.00043", "71336": "D2146.00044", "71340": "D2146.00045", 
            "71346": "D2146.00046", "71697": "D2146.00047", "73884": "D2146.00048", "75022": "D2146.00050", 
            "78215": "D2146.00051", "78628": "D2146.00052", "87858": "D2146.00054", "75022": "D2146.00055", 
            "95581": "D2146.00057", "53000": "D2575.00001", "64695": "D2575.00002", "65103": "D2575.00003", 
            "67985": "D2575.00004", "67983": "D2575.00006", "68360": "D2575.00007", "68544": "D2575.00008", 
            "68681": "D2575.00009", "68977": "D2575.00010", "68979": "D2575.00011", "68981": "D2575.00012", 
            "74389": "D2575.00015", "93894": "D2575.00016", "51180": "D1632.00001", "61972": "D1632.00002", 
            "63262": "D1632.00003", "69998": "D1632.00004", "71009": "D1632.00005", "80459": "D1632.00006", 
            "80858": "D1632.00007", "84342": "D1632.00008", "61409": "D1669.00001", "79134": "D1399.00001", 
            "79988": "D1399.00002", "82315": "D1399.00003", "82403": "D1399.00004", "83071": "D1399.00005", 
            "90226": "D1399.00006", "92343": "D1399.00007", "92785": "D1399.00008", "92861": "D1399.00010", 
            "96725": "D1399.00012", "96822": "D1399.00013", "75767": "D2731.00001", "75837": "D2731.00002", 
            "76113": "D2731.00003", "76877": "D2731.00004", "83123": "D2731.00005", "90868": "D2731.00006", 
            "93909": "D2731.00007", "93911": "D2731.00008", "93913": "D2731.00009"
        }


        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados Volantes"

        headers = ["Código Maestro", "Distribuidor", "Fecha de Consulta", "Tipo de Volante", "Mensaje", "Datos Volantes"]
        ws.append(headers)

        # Aplicar estilos a las cabeceras
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Aplicar bordes a las celdas
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        # Insertar los datos con formato
        for i, resultado in enumerate(self.resultados, start=2):
            distribuidor = resultado["Distribuidor"]
            codigo_maestro = resultado["Codigo"]
            distribuidor_mapeado = distribuidor_a_codigo.get(distribuidor, "No asignado")
            tipo_volante = "Voz" if resultado["Volantetype"] == "V" else "Reposición"
            
            # Concatenar los volantes si existen
            datos_volantes = ""
            if "Datos" in resultado:
                for volante in resultado["Datos"]:
                    volante_info = f"Secuencia: {volante['Secuencia']}, Efectivo: {volante['Valor Efectivo']}, Cheque: {volante['Valor Cheque']}"
                    datos_volantes += f"{volante_info}; "

            fila = [codigo_maestro, distribuidor_mapeado, resultado["Fecha"], tipo_volante, resultado.get("Mensaje", ""), datos_volantes]
            for col_num, cell_value in enumerate(fila, start=1):
                cell = ws.cell(row=i, column=col_num, value=cell_value)
                cell.border = thin_border
                if col_num == 6:  # Datos Volantes, que puede tener mucho texto
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Ajustar tamaño de las columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # Add a little extra padding
            ws.column_dimensions[column].width = adjusted_width

        # Guardar el archivo Excel con formato
        try:
            # Obtener la fecha actual
            fecha_actual = QDate.currentDate().toString("yyyy-MM-dd")
            archivo_excel = f"Consulta-Volantes-{fecha_actual}.xlsx"

            wb.save(archivo_excel)
            self.log_progreso(f"Resultados exportados a Excel en {archivo_excel}")
        except Exception as e:
            self.log_progreso(f"Error al exportar a Excel: {e}")

